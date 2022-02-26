#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Author: wp
# @Time: 2021/5/23 22:17
# @File: get_data.py


import openpyxl


class GetData:

    def __init__(self, filename):
        self.filename = filename

    # 读excel
    def read_excel(self):
        excel = openpyxl.load_workbook(self.filename)
        excel_sheet = excel.active
        data_list = []
        row_list = []
        rows = excel_sheet.max_row
        cols = excel_sheet.max_column
        for i in range(2, rows + 1):
            for j in range(1, cols + 1):
                value = excel_sheet.cell(i, j).value
                row_list.append(value)
            data_list.append(row_list)
            row_list = []
        return data_list

    # 写excel
    def write_excel(self, row, col, text):
        excel = openpyxl.load_workbook(self.filename)
        excel_sheet = excel.active
        # 行、列索引从1开始
        excel_sheet.cell(row, col).value = text
        excel.save(self.filename)


if __name__ == '__main__':
    data = GetData('./case.xlsx')
    print(data.read_excel())
    print(len(data.read_excel()))
    print(data.read_excel()[3])
    print("---------")
    print(data.read_excel()[3][9])
    print(type(data.read_excel()[3][9]))
