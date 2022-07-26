#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Author: wp
# @Time: 2021/5/23 22:17
# @File: excel_utils.py


import openpyxl
from openpyxl.styles import PatternFill, Font


class GetExcelData:

    def __init__(self, filename):
        self.filename = filename
        self.excel = openpyxl.load_workbook(self.filename)
        self.excel_sheet = self.excel.active

    # 读excel
    def read_excel(self):
        excel = openpyxl.load_workbook(self.filename)
        data_list = []
        sheets = excel.sheetnames
        for sheet1 in sheets:
            sheet = excel[sheet1]
            for values in sheet.values:
                if type(list(values)[0]) is int:
                    data_list.append(list(values))
        return data_list

    # 写excel
    def write_excel(self, row, col, text):
        # 行、列索引从1开始
        self.excel_sheet.cell(row, col).value = text
        self.excel.save(self.filename)

    # Pass配置
    def pass_(self, row, column):
        self.excel_sheet.cell(row=row, column=column).value = 'Pass'
        # 写入单元格样式设定：绿色+加粗
        self.excel_sheet.cell(row=row, column=column).fill = PatternFill('solid', fgColor='AACF91')
        self.excel_sheet.cell(row=row, column=column).font = Font(bold=True)
        self.excel.save(self.filename)

    # Failed配置
    def failed_(self, row, column):
        self.excel_sheet.cell(row=row, column=column).value = 'Failed'
        # 写入单元格样式设定：红色+加粗
        self.excel_sheet.cell(row=row, column=column).fill = PatternFill('solid', fgColor='FF0000')
        self.excel_sheet.cell(row=row, column=column).font = Font(bold=True)
        self.excel.save(self.filename)

    # Skip配置
    def skip_(self, row, column):
        self.excel_sheet.cell(row=row, column=column).value = 'Skip'
        # 写入单元格样式设定：黄色+加粗
        self.excel_sheet.cell(row=row, column=column).fill = PatternFill('solid', fgColor='FFFF00')
        self.excel_sheet.cell(row=row, column=column).font = Font(bold=True)
        self.excel.save(self.filename)


if __name__ == '__main__':
    data = GetExcelData('../../data/api_case.xlsx')
    print(data.read_excel())
    print(len(data.read_excel()))
    # print(data.read_excel()[3])
    # print("---------")
    # print(data.read_excel()[3][9])
    # print(type(data.read_excel()[3][9]))
